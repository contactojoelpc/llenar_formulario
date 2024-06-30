# import pandas as pd

# Función para leer un archivo Excel y procesar los registros
# def procesar_excel(file_path):
#     Leer el archivo Excel
#     df = pd.read_excel(file_path)
    
#     Iterar sobre cada fila del DataFrame
#     for index, row in df.iterrows():
#         Crear el diccionario con las llaves especificadas
#         registro = {
#             'campo1': row['campo1'],
#             'campo2': row['campo2'],
#             'campo3': row['campo3']
#         }
#         Imprimir el diccionario
#         print(registro)

# Ruta al archivo Excel
# file_path = 'ruta/al/archivo.xlsx'

# Llamada a la función para procesar el archivo Excel
# procesar_excel(file_path)









# import openpyxl

# def leer_excel(nombre_archivo):
#     # Cargar el archivo Excel
#     wb = openpyxl.load_workbook(nombre_archivo)
#     sheet = wb.active  # Seleccionar la hoja activa

#     registros = []
    
#     # Iterar sobre las filas del archivo Excel
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         # Crear un diccionario para cada registro
#         registro = {
#             'campo1': row[0],
#             'campo2': row[1],
#             'campo3': row[2]
#         }
#         registros.append(registro)
        
#         # Imprimir el registro actual
#         print(f'Registro actual: {registro}')
        
#     return registros

# # Ejemplo de uso
# nombre_archivo = 'ejemplo.xlsx'  # Reemplazar con el nombre de tu archivo Excel
# registros = leer_excel(nombre_archivo)












# import openpyxl

# def leer_excel(nombre_archivo):
#     Cargar el archivo Excel
#     wb = openpyxl.load_workbook(nombre_archivo)
#     sheet = wb.active  # Seleccionar la hoja activa

#     registros = []
    
#     Iterar sobre las filas del archivo Excel comenzando desde la fila 2 y columna B
#     for row in sheet.iter_rows(min_row=2, min_col=2, values_only=True):
#         Comprobar si la fila está vacía
#         if all(celda is None for celda in row):
#             break
        
#         Crear un diccionario para cada registro solo si no hay celdas vacías
#         if all(celda is not None for celda in row[:3]):
#             registro = {
#                 'campo1': row[0],
#                 'campo2': row[1],
#                 'campo3': row[2]
#             }
#             registros.append(registro)
            
#             Imprimir el registro actual
#             print(f'Registro actual: {registro}')
        
#     return registros

# Ejemplo de uso
# nombre_archivo = 'ejemplo.xlsx'  # Reemplazar con el nombre de tu archivo Excel
# registros = leer_excel(nombre_archivo)











# ****************************** SI SIRVE ******************************

# import openpyxl

# def leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B'):
#     # Cargar el archivo Excel
#     wb = openpyxl.load_workbook(nombre_archivo)
#     sheet = wb.active  # Seleccionar la hoja activa

#     registros = []
    
#     # Convertir la letra de la columna de inicio a índice de columna
#     columna_inicio_idx = openpyxl.utils.column_index_from_string(columna_inicio)
    
#     # Iterar sobre las filas del archivo Excel comenzando desde la fila y columna especificadas
#     for row in sheet.iter_rows(min_row=fila_inicio, min_col=columna_inicio_idx, values_only=True):
#         # Comprobar si la fila está vacía
#         if all(celda is None for celda in row):
#             break
        
#         # Crear un diccionario para cada registro solo si no hay celdas vacías
#         if all(celda is not None for celda in row[:3]):
#             registro = {
#                 'campo1': row[0],
#                 'campo2': row[1],
#                 'campo3': row[2]
#             }
#             registros.append(registro)
            
#             # Imprimir el registro actual
#             print(f'Registro actual: {registro}')
        
#     return registros

# # Ejemplo de uso
# nombre_archivo = 'ejemplo.xlsx'  # Reemplazar con el nombre de tu archivo Excel
# registros = leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B')








# ****************************** PRUEBA CON SELENIUM SI SIRVE ******************************

# import openpyxl
# import time
# from selenium import webdriver
# from selenium.webdriver.common.by import By

# def leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B'):
#     # Cargar el archivo Excel
#     wb = openpyxl.load_workbook(nombre_archivo)
#     sheet = wb.active  # Seleccionar la hoja activa

#     registros = []
    
#     # Convertir la letra de la columna de inicio a índice de columna
#     columna_inicio_idx = openpyxl.utils.column_index_from_string(columna_inicio)
    
#     # Iterar sobre las filas del archivo Excel comenzando desde la fila y columna especificadas
#     for row in sheet.iter_rows(min_row=fila_inicio, min_col=columna_inicio_idx, values_only=True):
#         # Comprobar si la fila está vacía
#         if all(celda is None for celda in row):
#             break
        
#         # Crear un diccionario para cada registro solo si no hay celdas vacías
#         if all(celda is not None for celda in row[:3]):
#             registro = {
#                 'campo1': row[0],
#                 'campo2': row[1],
#                 'campo3': row[2]
#             }
#             registros.append(registro)
            
#             # Imprimir el registro actual
#             print(f'Registro actual: {registro}')
        
#     return registros

# def llenar_formulario(registros):
#     # Configurar Selenium y abrir el navegador
#     driver = webdriver.Chrome()  # Asegúrate de que el ChromeDriver esté en tu PATH o especifica la ruta completa

#     # Reemplazar con la URL de la página que contiene el formulario
#     url = 'file:///Users/joelpalma/Documents/ejercicios_python/birmex_python/formulario.html'
#     driver.get(url)
    
#     for registro in registros:
#         # Llenar los inputs del formulario con los valores del registro
#         driver.find_element(By.ID, 'campo1').send_keys(registro['campo1'])
#         driver.find_element(By.ID, 'campo2').send_keys(registro['campo2'])
#         driver.find_element(By.ID, 'campo3').send_keys(registro['campo3'])
        
#         # Presionar el botón con id 'btnenter'
#         driver.find_element(By.ID, 'btnenter').click()
        
#         # Esperar 3 segundos antes de procesar el siguiente registro
#         time.sleep(3)
        
#         # Limpiar los inputs del formulario si es necesario
#         driver.find_element(By.ID, 'campo1').clear()
#         driver.find_element(By.ID, 'campo2').clear()
#         driver.find_element(By.ID, 'campo3').clear()
    
#     driver.quit()

# # Ejemplo de uso
# nombre_archivo = 'ejemplo.xlsx'  # Reemplazar con el nombre de tu archivo Excel
# registros = leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B')
# llenar_formulario(registros)







# ****************************** PRUEBA1 CON SELENIUM SI SIRVE ******************************

# import openpyxl
# import time
# from selenium import webdriver
# from selenium.webdriver.common.by import By

# def leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B'):
#     # Cargar el archivo Excel
#     wb = openpyxl.load_workbook(nombre_archivo)
#     sheet = wb.active  # Seleccionar la hoja activa

#     registros = []
    
#     # Convertir la letra de la columna de inicio a índice de columna
#     columna_inicio_idx = openpyxl.utils.column_index_from_string(columna_inicio)
    
#     # Iterar sobre las filas del archivo Excel comenzando desde la fila y columna especificadas
#     for row in sheet.iter_rows(min_row=fila_inicio, min_col=columna_inicio_idx, values_only=True):
#         # Comprobar si la fila está vacía
#         if all(celda is None for celda in row):
#             break
        
#         # Crear un diccionario para cada registro solo si no hay celdas vacías
#         if all(celda is not None for celda in row[:3]):
#             registro = {
#                 'campo1': row[0],
#                 'campo2': row[1],
#                 'campo3': row[2]
#             }
#             registros.append(registro)
            
#             # Imprimir el registro actual
#             print(f'Registro actual: {registro}')
        
#     return registros

# def llenar_formulario(registros):
#     # Configurar Selenium y abrir el navegador
#     driver = webdriver.Chrome()  # Asegúrate de que el ChromeDriver esté en tu PATH o especifica la ruta completa

#     # Reemplazar con la URL de la página que contiene el formulario
#     url = 'file:///Users/joelpalma/Documents/ejercicios_python/birmex_python/formulario.html'
#     driver.get(url)
    
#     for i, registro in enumerate(registros):
#         # Llenar los inputs del formulario con los valores del registro
#         driver.find_element(By.ID, 'campo1').send_keys(registro['campo1'])
#         driver.find_element(By.ID, 'campo2').send_keys(registro['campo2'])
#         driver.find_element(By.ID, 'campo3').send_keys(registro['campo3'])
        
#         # Presionar el botón con id 'btnenter'
#         driver.find_element(By.ID, 'btnenter').click()
        
#         # Imprimir mensaje después de presionar el botón
#         print(f'Registro {i+1} enviado: {registro}')
        
#         # Esperar 3 segundos antes de procesar el siguiente registro
#         time.sleep(3)
        
#         # Limpiar los inputs del formulario si es necesario
#         driver.find_element(By.ID, 'campo1').clear()
#         driver.find_element(By.ID, 'campo2').clear()
#         driver.find_element(By.ID, 'campo3').clear()
    
#     driver.quit()

# # Ejemplo de uso
# nombre_archivo = 'ejemplo.xlsx'  # Reemplazar con el nombre de tu archivo Excel
# registros = leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B')
# llenar_formulario(registros)






# ****************************** PRUEBA2 CON SELENIUM SI SIRVE ******************************

import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By

def leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B'):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(nombre_archivo)
    sheet = wb.active  # Seleccionar la hoja activa

    registros = []
    
    # Convertir la letra de la columna de inicio a índice de columna
    columna_inicio_idx = openpyxl.utils.column_index_from_string(columna_inicio)
    
    # Iterar sobre las filas del archivo Excel comenzando desde la fila y columna especificadas
    for row in sheet.iter_rows(min_row=fila_inicio, min_col=columna_inicio_idx, values_only=True):
        # Comprobar si la fila está vacía
        if all(celda is None for celda in row):
            break
        
        # Crear un diccionario para cada registro solo si no hay celdas vacías
        if all(celda is not None for celda in row[:4]):
            registro = {
                'campo1': row[0],
                'campo2': row[1],
                'campo3': row[2],
                'campo4': row[3]
            }
            registros.append(registro)
        
    return registros

def llenar_formulario(registros):
    # Configurar Selenium y abrir el navegador
    driver = webdriver.Chrome()  # Asegúrate de que el ChromeDriver esté en tu PATH o especifica la ruta completa

    # Reemplazar con la URL de la página que contiene el formulario
    url = 'file:///Users/joelpalma/Documents/ejercicios_python/birmex_python/formulario.html'
    driver.get(url)
    
    for i, registro in enumerate(registros):
        # Llenar los inputs del formulario con los valores del registro
        driver.find_element(By.ID, 'campo1').send_keys(registro['campo1'])
        driver.find_element(By.ID, 'campo2').send_keys(registro['campo2'])
        driver.find_element(By.ID, 'campo3').send_keys(registro['campo3'])
        driver.find_element(By.ID, 'campo4').send_keys(registro['campo4'])
        
        # Presionar el botón con id 'btnenter'
        driver.find_element(By.ID, 'btnenter').click()
        
        # Imprimir mensaje después de presionar el botón
        print(f'Registro {i+1} enviado: {registro}')
        
        # Esperar 3 segundos antes de procesar el siguiente registro
        time.sleep(3)
        
        # Si no es el último registro, espera a que el formulario se recargue
        if i < len(registros) - 1:
            # Limpiar los inputs del formulario
            driver.find_element(By.ID, 'campo1').clear()
            driver.find_element(By.ID, 'campo2').clear()
            driver.find_element(By.ID, 'campo3').clear()
            # driver.find_element(By.ID, 'campo4').clear()
    
    driver.quit()

# Ejemplo de uso
nombre_archivo = 'ejemplo.xlsx'  # Reemplazar con el nombre de tu archivo Excel
registros = leer_excel(nombre_archivo, fila_inicio=2, columna_inicio='B')
llenar_formulario(registros)
